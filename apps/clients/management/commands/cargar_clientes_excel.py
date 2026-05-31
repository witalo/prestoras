"""
Comando Django para cargar clientes desde el Excel LISTA DE CLIENTES 1222.xlsx
Uso: python manage.py cargar_clientes_excel [ruta_excel] [--company-id N] [--dry-run]
"""
import re
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.companies.models import Company
from apps.zones.models import Zone
from apps.clients.models import Client


# ─── Normalización de nombres de zona ────────────────────────────────────────
# Unifica variantes tipográficas del mismo distrito/ruta en un solo nombre canónico.
ZONA_NORMALIZAR = {
    # Mariano Melgar
    'M. MELGAR': 'MARIANO MELGAR',
    # Cercado variantes
    'CERCAD PARTE 2':    'CERCADO PARTE 2',
    'CERCADO PARTE3':    'CERCADO PARTE 3',
    'CERCADO( SEMANAL)': 'CERCADO (SEMANAL)',
    'CERCADO(SEMANAL)':  'CERCADO (SEMANAL)',
    'CERCADO(PARALELO)': 'CERCADO (PARALELO)',
    # JLBR variantes
    'J.L.BYR.':          'JLBR',
    'JLBYR(PARALELO':    'JLBR (PARALELO)',
    'JLBR(SEMANAL)':     'JLBR (SEMANAL)',
    # Jacobo Hunter
    'HANTER':            'JACOBO HUNTER',
    'JACOBO HANTER':     'JACOBO HUNTER',
    # Miraflores
    'MIRAFLORES(PARALELO)': 'MIRAFLORES (PARALELO)',
    'MIRAFLORES (PARALELO)': 'MIRAFLORES (PARALELO)',
    'MIRAFLORES (SEMANAL)':  'MIRAFLORES (SEMANAL)',
    # Socabaya (corrige Socobaya)
    'SOCOBAYA (PARALELO)': 'SOCABAYA (PARALELO)',
    'SOCABAYA(SEMANAL)':   'SOCABAYA (SEMANAL)',
    # Paucarpata
    'PAUCARPATA(PARALELO)': 'PAUCARPATA (PARALELO)',
    'PAUCARPATA(SEMANAL)':  'PAUCARPATA (SEMANAL)',
    # Otros
    'CONO NORTE(PARALELO)': 'CONO NORTE (PARALELO)',
    'ASA(PARALELO)':        'ASA (PARALELO)',
    'MARIANO MELGAR (PARALELO)': 'MARIANO MELGAR (PARALELO)',
    'MARIANO MELGAR (SEMANAL)':  'MARIANO MELGAR (SEMANAL)',
    'SABANDIA (SEMANAL)':        'SABANDIA (SEMANAL)',
}


def normalizar_zona(valor):
    """Limpia y unifica el nombre de zona."""
    if not valor:
        return None
    nombre = str(valor).strip()
    if not nombre:
        return None
    return ZONA_NORMALIZAR.get(nombre, nombre)


def extraer_dni(valor):
    """
    Extrae un DNI de 8 dígitos del valor de la celda.
    Maneja enteros, flotantes y strings con texto adicional (ej. 'OJO 02431024').
    """
    if valor is None:
        return None
    if isinstance(valor, float):
        valor = int(valor)
    if isinstance(valor, int):
        dni = str(valor).zfill(8)
        return dni if len(dni) == 8 and dni.isdigit() else None
    # String: buscar primera secuencia de 8 dígitos
    s = str(valor).strip()
    matches = re.findall(r'\b\d{8}\b', s)
    if matches:
        return matches[0]
    return None


def limpiar_nombre(valor):
    """Limpia y capitaliza nombres."""
    if not valor:
        return ''
    return ' '.join(str(valor).strip().split())  # elimina espacios múltiples


def limpiar_telefono(valor):
    """Convierte teléfono a string limpio."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        t = str(int(valor))
        return t if len(t) >= 7 else None
    t = str(valor).strip()
    # Quitar caracteres no numéricos excepto +
    t = re.sub(r'[^\d+]', '', t)
    return t if len(t) >= 7 else None


def limpiar_email(valor):
    """Valida formato básico de email."""
    if not valor:
        return None
    e = str(valor).strip().lower()
    return e if '@' in e and '.' in e.split('@')[-1] else None


class Command(BaseCommand):
    help = 'Carga zonas y clientes desde Excel. Evita duplicados.'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_path',
            nargs='?',
            default=r'C:\Users\User\Downloads\Nueva carpeta (2)\LISTA DE CLIENTES 1222.xlsx',
            help='Ruta al archivo Excel (por defecto: ruta del archivo original)'
        )
        parser.add_argument(
            '--company-id',
            type=int,
            default=None,
            help='ID de la empresa donde se cargarán los clientes (por defecto: primera activa)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='Simular sin guardar nada en la base de datos'
        )

    def handle(self, *args, **options):
        excel_path = options['excel_path']
        company_id = options.get('company_id')
        dry_run    = options['dry_run']

        if dry_run:
            self.stdout.write(self.style.WARNING('*** MODO DRY-RUN: no se guardará nada ***\n'))

        # ── Obtener empresa ──────────────────────────────────────────────────
        if company_id:
            try:
                company = Company.objects.get(id=company_id)
            except Company.DoesNotExist:
                self.stderr.write(self.style.ERROR(f'No existe empresa con ID={company_id}'))
                return
        else:
            company = Company.objects.filter(is_active=True).first()
            if not company:
                self.stderr.write(self.style.ERROR('No hay empresas activas en la base de datos.'))
                return

        self.stdout.write(
            f'Empresa: {self.style.SUCCESS(company.commercial_name or company.legal_name)} '
            f'(ID={company.id})\n'
        )

        # ── Leer Excel ───────────────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except FileNotFoundError:
            self.stderr.write(self.style.ERROR(f'Archivo no encontrado: {excel_path}'))
            return

        ws = wb.active

        # Recolectar filas que tengan al menos nombre o DNI
        filas_validas = []
        for fila in range(2, ws.max_row + 1):
            vals = [ws.cell(fila, c).value for c in range(1, 8)]
            if vals[0] or vals[2]:   # tiene nombre o DNI
                filas_validas.append(vals)

        self.stdout.write(f'Filas con datos en el Excel: {len(filas_validas)}\n')

        # ── Paso 1: Zonas únicas ─────────────────────────────────────────────
        nombres_zona = set()
        for r in filas_validas:
            z = normalizar_zona(r[4])
            if z:
                nombres_zona.add(z)

        self.stdout.write(f'Zonas únicas a cargar: {len(nombres_zona)}')

        zona_cache = {}          # nombre → instancia Zone
        zonas_creadas    = 0
        zonas_existentes = 0

        with transaction.atomic():
            for nombre in sorted(nombres_zona):
                if dry_run:
                    existe = Zone.objects.filter(company=company, name=nombre).exists()
                    if existe:
                        zonas_existentes += 1
                        self.stdout.write(f'  ~ Zona ya existe:  {nombre}')
                    else:
                        zonas_creadas += 1
                        self.stdout.write(f'  + Zona nueva (dry): {nombre}')
                    zona_cache[nombre] = None
                else:
                    zona, created = Zone.objects.get_or_create(
                        company=company,
                        name=nombre,
                        defaults={'status': 'ACTIVE'}
                    )
                    zona_cache[nombre] = zona
                    if created:
                        zonas_creadas += 1
                        self.stdout.write(f'  {self.style.SUCCESS("+")} Zona creada:   {nombre}')
                    else:
                        zonas_existentes += 1
                        self.stdout.write(f'  ~ Zona existente: {nombre}')

        # ── Paso 2: Clientes ─────────────────────────────────────────────────
        self.stdout.write(f'\nCargando clientes...')

        clientes_creados    = 0
        clientes_existentes = 0
        clientes_sin_dni    = 0
        clientes_error      = 0
        dnis_vistos         = set()   # para detectar duplicados dentro del propio Excel

        with transaction.atomic():
            for r in filas_validas:
                first_name = limpiar_nombre(r[0])
                last_name  = limpiar_nombre(r[1])
                dni        = extraer_dni(r[2])
                address    = limpiar_nombre(r[3])
                zona_nombre = normalizar_zona(r[4])
                phone      = limpiar_telefono(r[5])
                email      = limpiar_email(r[6])

                # Saltar si no tiene DNI
                if not dni:
                    clientes_sin_dni += 1
                    self.stdout.write(
                        self.style.WARNING(f'  SKIP sin DNI: {first_name} {last_name}')
                    )
                    continue

                # Detectar duplicado en el mismo Excel
                if dni in dnis_vistos:
                    self.stdout.write(
                        self.style.WARNING(f'  SKIP duplicado en Excel: DNI {dni} ({first_name} {last_name})')
                    )
                    continue
                dnis_vistos.add(dni)

                zona = zona_cache.get(zona_nombre) if zona_nombre else None

                if dry_run:
                    existe = Client.objects.filter(dni=dni).exists()
                    if existe:
                        clientes_existentes += 1
                    else:
                        clientes_creados += 1
                    continue

                try:
                    client, created = Client.objects.get_or_create(
                        dni=dni,
                        defaults={
                            'company':       company,
                            'first_name':    first_name or 'SIN NOMBRE',
                            'last_name':     last_name  or 'SIN APELLIDO',
                            'phone':         phone,
                            'email':         email,
                            'home_address':  address or None,
                            'zone':          zona,
                            'classification':'REGULAR',
                            'is_active':     True,
                        }
                    )
                    if created:
                        clientes_creados += 1
                    else:
                        clientes_existentes += 1
                except Exception as e:
                    clientes_error += 1
                    self.stdout.write(
                        self.style.ERROR(f'  ERROR DNI {dni} ({first_name} {last_name}): {e}')
                    )

        # ── Resumen ──────────────────────────────────────────────────────────
        sep = '=' * 55
        self.stdout.write(f'\n{sep}')
        self.stdout.write(
            self.style.SUCCESS(
                f'ZONAS:    {zonas_creadas:>3} creadas  |  {zonas_existentes:>3} ya existían'
            )
        )
        self.stdout.write(
            self.style.SUCCESS(
                f'CLIENTES: {clientes_creados:>3} creados  |  {clientes_existentes:>3} ya existían  '
                f'|  {clientes_sin_dni} sin DNI  |  {clientes_error} errores'
            )
        )
        if dry_run:
            self.stdout.write(self.style.WARNING('\n*** Dry-run: ningún cambio fue guardado ***'))
        self.stdout.write(sep)
